"""
Import Router - Import Chart of Accounts and Voucher History from Excel files
Supports Lebanese LCOA format with supplier/client detection
"""
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File, Form
from typing import Optional
import uuid
import json
from datetime import datetime, timezone
from collections import defaultdict
import io
import logging

from core.database import db
from core.auth import get_current_user

router = APIRouter(tags=["Import"])
logger = logging.getLogger(__name__)


# ================== PREVIEW HEADERS ==================

@router.post("/import/preview-headers")
async def preview_excel_headers(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Read Excel file headers and first few rows for field mapping preview"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not available")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active
    
    rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    wb.close()
    
    if not rows:
        raise HTTPException(status_code=400, detail="Empty file")
    
    headers = [str(h) if h else f'Column {i}' for i, h in enumerate(rows[0])]
    
    # Get sample data from first 3 rows
    sample_rows = []
    for row in rows[1:4]:
        sample_rows.append([str(v) if v is not None else '' for v in row])
    
    return {
        "headers": headers,
        "header_count": len(headers),
        "sample_rows": sample_rows,
        "filename": file.filename
    }

# Account type mapping based on LCOA class
ACCOUNT_CLASS_TYPES = {
    1: 'equity',      # Capital
    2: 'asset',       # Fixed Assets
    3: 'asset',       # Inventory
    4: 'liability',   # Receivables/Payables
    5: 'asset',       # Financial accounts
    6: 'expense',     # Expenses
    7: 'revenue',     # Revenue
    8: 'expense',     # Special results
}


@router.post("/import/chart-of-accounts")
async def import_chart_of_accounts(
    file: UploadFile = File(...),
    organization_id: str = Form(...),
    fiscal_year_id: str = Form(None),
    field_mapping: str = Form(None),  # JSON: {"account_code": 0, "account_name": 2, "address": 14, "phone": 15, "reg_id": 18, "regno": 27}
    current_user: dict = Depends(get_current_user)
):
    """
    Import Chart of Accounts from Excel file with optional field mapping.
    If field_mapping is provided, uses custom column indexes.
    Otherwise uses default LCOA format.
    """
    if current_user['role'] not in ['super_admin', 'admin']:
        raise HTTPException(status_code=403, detail="Only admins can import data")
    
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl library not available")
    
    # Read the file
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active
    
    # Parse field mapping
    mapping = None
    if field_mapping:
        try:
            mapping = json.loads(field_mapping)
        except json.JSONDecodeError:
            pass
    
    # Helper to get value from row using mapping or default index
    def get_col(row, field_name, default_idx):
        idx = mapping.get(field_name, default_idx) if mapping else default_idx
        if idx is None or idx < 0 or idx == '':
            return ''
        idx = int(idx)
        return row[idx] if len(row) > idx and row[idx] is not None else ''
    
    accounts_created = 0
    accounts_updated = 0
    suppliers_created = 0
    customers_created = 0
    errors = []
    
    # Track unique codes to avoid duplicates in the file
    seen_codes = set()
    
    accounts_to_insert = []
    
    row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1
        try:
            code = str(get_col(row, 'account_code', 0)).strip()
            if not code or not code[0].isdigit():
                continue
            
            # Skip duplicate codes in file
            if code in seen_codes:
                continue
            seen_codes.add(code)
            
            account_name = str(get_col(row, 'account_name', 2)).strip()
            account_type_ar = str(get_col(row, 'account_type', 3)).strip()
            
            # Determine account class from first digit
            account_class = int(code[0]) if code[0].isdigit() else None
            account_type = ACCOUNT_CLASS_TYPES.get(account_class, 'asset')
            
            # Special handling for class 4 sub-types
            if code.startswith('40'):
                account_type = 'liability'
            elif code.startswith('41'):
                account_type = 'asset'
            elif code.startswith('44'):
                account_type = 'liability'
            elif code.startswith('45'):
                account_type = 'liability'
            
            # Contact info for suppliers/customers
            name_field = str(get_col(row, 'contact_name', 13)).strip()
            address_field = str(get_col(row, 'address', 14)).strip()
            phone_field = str(get_col(row, 'phone', 15)).strip()
            regno_field = str(get_col(row, 'regno', 27)).strip()
            
            account_doc = {
                'id': str(uuid.uuid4()),
                'code': code,
                'name': account_name,
                'name_ar': account_name,
                'account_class': account_class,
                'account_type': account_type,
                'organization_id': organization_id,
                'balance_lbp': 0,
                'balance_usd': 0,
                'is_active': True,
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            
            # Add contact info for suppliers and customers
            is_supplier = code.startswith('40') and len(code) > 4
            is_customer = code.startswith('41') and len(code) > 4
            
            # Region ID for customers (col 18 = REG_ID)
            region_id = str(get_col(row, 'reg_id', 18)).strip()
            
            if is_supplier or is_customer:
                # Store contact fields as TOP-LEVEL fields (matching existing app schema)
                account_doc['mobile'] = phone_field
                account_doc['address'] = address_field
                account_doc['contact_person'] = name_field or account_name
                if regno_field and regno_field != '0':
                    account_doc['registration_number'] = regno_field
                if is_customer and region_id and region_id != '0':
                    account_doc['region_id'] = region_id
                # Also keep contact_info for backward compat
                account_doc['contact_info'] = {
                    'name': name_field or account_name,
                    'address': address_field,
                    'phone': phone_field,
                    'is_supplier': is_supplier,
                    'is_customer': is_customer,
                    'registration_number': regno_field if (regno_field and regno_field != '0') else '',
                    'region_id': region_id if (is_customer and region_id and region_id != '0') else ''
                }
            
            accounts_to_insert.append(account_doc)
            
        except Exception as e:
            errors.append(f"Row {row_num + 1}: {str(e)}")
            if len(errors) > 50:
                break
    
    wb.close()
    
    # Batch insert/update accounts
    if accounts_to_insert:
        # Check for existing accounts by code
        existing_codes = set()
        existing_accounts = await db.accounts.find(
            {'organization_id': organization_id, 'code': {'$in': [a['code'] for a in accounts_to_insert]}},
            {'code': 1, '_id': 0}
        ).to_list(None)
        existing_codes = {a['code'] for a in existing_accounts}
        
        new_accounts = []
        for acc in accounts_to_insert:
            if acc['code'] in existing_codes:
                # Update existing account
                update_data = {
                    'name': acc['name'],
                    'name_ar': acc['name_ar'],
                    'account_class': acc['account_class'],
                    'account_type': acc['account_type'],
                }
                if acc.get('contact_info'):
                    update_data['contact_info'] = acc['contact_info']
                # Also update top-level contact fields
                for field in ['mobile', 'address', 'contact_person', 'registration_number', 'region_id']:
                    if field in acc:
                        update_data[field] = acc[field]
                await db.accounts.update_one(
                    {'code': acc['code'], 'organization_id': organization_id},
                    {'$set': update_data}
                )
                accounts_updated += 1
            else:
                new_accounts.append(acc)
                accounts_created += 1
                
                # Count suppliers/customers
                if acc.get('contact_info', {}).get('is_supplier'):
                    suppliers_created += 1
                if acc.get('contact_info', {}).get('is_customer'):
                    customers_created += 1
        
        # Batch insert new accounts
        if new_accounts:
            # Insert in batches of 500
            for i in range(0, len(new_accounts), 500):
                batch = new_accounts[i:i+500]
                await db.accounts.insert_many(batch)
    
    return {
        "message": "Chart of Accounts imported successfully",
        "accounts_created": accounts_created,
        "accounts_updated": accounts_updated,
        "suppliers_detected": suppliers_created,
        "customers_detected": customers_created,
        "total_processed": len(accounts_to_insert),
        "errors": errors[:20],  # Return first 20 errors
        "error_count": len(errors)
    }


@router.post("/import/vouchers")
async def import_vouchers(
    file: UploadFile = File(...),
    organization_id: str = Form(...),
    fiscal_year_id: str = Form(None),
    field_mapping: str = Form(None),  # JSON: {"tran": 0, "account_code": 3, "date": 5, "cr_lbp": 8, "dr_lbp": 10, "cr_usd": 11, "dr_usd": 12, "description": 14, "currency": 17}
    current_user: dict = Depends(get_current_user)
):
    """Import voucher transactions with optional field mapping."""
    if current_user['role'] not in ['super_admin', 'admin']:
        raise HTTPException(status_code=403, detail="Only admins can import data")
    
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl library not available")
    
    # Parse field mapping
    mapping = None
    if field_mapping:
        try:
            mapping = json.loads(field_mapping)
        except json.JSONDecodeError:
            pass
    
    def get_col(row, field_name, default_idx):
        idx = mapping.get(field_name, default_idx) if mapping else default_idx
        if idx is None or idx < 0 or idx == '':
            return None
        idx = int(idx)
        return row[idx] if len(row) > idx else None
    
    # Read the file
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active
    
    # Step 1: Group all rows by TRAN (voucher ID)
    voucher_groups = defaultdict(list)
    row_count = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1
        tran_id = get_col(row, 'tran', 0)
        if tran_id is None:
            continue
        voucher_groups[tran_id].append(row)
    
    wb.close()
    
    logger.info(f"Import: {row_count} rows, {len(voucher_groups)} vouchers to process")
    
    # Step 1b: Load fiscal year date range if specified
    fy_start = None
    fy_end = None
    if fiscal_year_id and fiscal_year_id != 'none':
        fy = await db.fiscal_years.find_one({'id': fiscal_year_id}, {'_id': 0})
        if fy:
            fy_start = fy['start_date']
            fy_end = fy['end_date']
            logger.info(f"FY filter active: {fy['name']} ({fy_start} to {fy_end})")
    
    # Step 2: Get current voucher sequence number
    last_voucher = await db.vouchers.find(
        {'organization_id': organization_id, 'voucher_type': 'JV'},
        {'voucher_number': 1}
    ).sort('created_at', -1).limit(1).to_list(1)
    
    # Generate sequence starting point
    seq_num = 1
    if last_voucher:
        try:
            last_num = last_voucher[0].get('voucher_number', 'JV-0000-00000')
            parts = last_num.split('-')
            seq_num = int(parts[-1]) + 1 if len(parts) > 1 else 1
        except (ValueError, IndexError):
            seq_num = 1
    
    vouchers_created = 0
    vouchers_failed = 0
    vouchers_skipped = 0
    lines_processed = 0
    errors = []
    
    # Step 3: Process each voucher group
    voucher_docs = []
    
    for tran_id, rows in voucher_groups.items():
        try:
            # Parse date from first row
            date_raw = str(get_col(rows[0], 'date', 5) or '')
            if len(date_raw) == 8:  # YYYYMMDD
                date_str = f"{date_raw[:4]}-{date_raw[4:6]}-{date_raw[6:8]}"
            elif len(date_raw) == 10:  # YYYY-MM-DD already
                date_str = date_raw
            else:
                date_str = '2016-01-01'  # Fallback
            
            # Skip vouchers outside the selected fiscal year
            if fy_start and fy_end:
                if date_str < fy_start or date_str > fy_end:
                    vouchers_skipped += 1
                    continue
            
            # Get description from first line
            desc_val = get_col(rows[0], 'description', 14)
            description = str(desc_val).strip() if desc_val else f'Import TRAN-{tran_id}'
            
            voucher_lines = []
            total_debit_lbp = 0
            total_credit_lbp = 0
            total_debit_usd = 0
            total_credit_usd = 0
            
            for line_row in rows:
                acct_val = get_col(line_row, 'account_code', 3)
                account_code = str(acct_val).strip() if acct_val else ''
                if not account_code:
                    continue
                
                # Extract amounts
                cr_lbp = float(get_col(line_row, 'cr_lbp', 8) or 0)
                dr_lbp = float(get_col(line_row, 'dr_lbp', 10) or 0)
                cr_usd = float(get_col(line_row, 'cr_usd', 11) or 0)
                dr_usd = float(get_col(line_row, 'dr_usd', 12) or 0)
                
                desc_v = get_col(line_row, 'description', 14)
                line_desc = str(desc_v).strip() if desc_v else ''
                cur_v = get_col(line_row, 'currency', 17)
                cur = str(cur_v) if cur_v else '1'
                
                # Exchange rate - 89500 for USD since 2023, 1507.5 before
                if cur == '2':
                    year = int(date_str[:4]) if len(date_str) >= 4 else 2016
                    exchange_rate = 89500.0 if year >= 2023 else 1507.5
                else:
                    exchange_rate = 1.0
                
                voucher_line = {
                    'account_code': account_code,
                    'description': line_desc,
                    'debit_lbp': dr_lbp,
                    'credit_lbp': cr_lbp,
                    'debit_usd': dr_usd,
                    'credit_usd': cr_usd,
                    'exchange_rate': exchange_rate
                }
                
                voucher_lines.append(voucher_line)
                total_debit_lbp += dr_lbp
                total_credit_lbp += cr_lbp
                total_debit_usd += dr_usd
                total_credit_usd += cr_usd
                lines_processed += 1
            
            if not voucher_lines:
                continue
            
            # Generate voucher number
            year = date_str[:4]
            voucher_number = f"JV-{year}-{seq_num:05d}"
            seq_num += 1
            
            voucher_doc = {
                'id': str(uuid.uuid4()),
                'voucher_number': voucher_number,
                'voucher_type': 'JV',
                'date': date_str,
                'reference': f'IMPORT-{tran_id}',
                'description': description,
                'lines': voucher_lines,
                'total_debit_lbp': total_debit_lbp,
                'total_credit_lbp': total_credit_lbp,
                'total_debit_usd': total_debit_usd,
                'total_credit_usd': total_credit_usd,
                'is_posted': True,
                'status': 'posted',
                'posted_at': datetime.now(timezone.utc).isoformat(),
                'organization_id': organization_id,
                'source_type': 'excel_import',
                'source_id': str(tran_id),
                'created_by': current_user['id'],
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            
            voucher_docs.append(voucher_doc)
            vouchers_created += 1
            
        except Exception as e:
            vouchers_failed += 1
            errors.append(f"TRAN {tran_id}: {str(e)}")
            if len(errors) > 50:
                break
    
    # Step 4: Batch insert vouchers (skip existing by source_id)
    vouchers_duplicate = 0
    if voucher_docs:
        # Check for existing vouchers by source_id to avoid duplicates on reimport
        existing_source_ids = set()
        existing = await db.vouchers.find(
            {'organization_id': organization_id, 'source_id': {'$in': [v['source_id'] for v in voucher_docs]}},
            {'source_id': 1, '_id': 0}
        ).to_list(None)
        existing_source_ids = {e['source_id'] for e in existing}
        
        new_vouchers = [v for v in voucher_docs if v['source_id'] not in existing_source_ids]
        vouchers_duplicate = len(voucher_docs) - len(new_vouchers)
        
        if new_vouchers:
            for i in range(0, len(new_vouchers), 500):
                batch = new_vouchers[i:i+500]
                await db.vouchers.insert_many(batch)
    else:
        new_vouchers = []
    
    # Step 5: Update account balances from NEW posted vouchers only
    balance_updates = defaultdict(lambda: {'lbp': 0, 'usd': 0})
    for voucher in new_vouchers:
        for line in voucher['lines']:
            code = line['account_code']
            balance_updates[code]['lbp'] += line['debit_lbp'] - line['credit_lbp']
            balance_updates[code]['usd'] += line['debit_usd'] - line['credit_usd']
    
    accounts_updated = 0
    for code, deltas in balance_updates.items():
        if deltas['lbp'] != 0 or deltas['usd'] != 0:
            result = await db.accounts.update_one(
                {'code': code, 'organization_id': organization_id},
                {'$inc': {'balance_lbp': deltas['lbp'], 'balance_usd': deltas['usd']}}
            )
            if result.modified_count > 0:
                accounts_updated += 1
    
    return {
        "message": "Voucher history imported successfully",
        "vouchers_created": len(new_vouchers),
        "vouchers_skipped": vouchers_skipped,
        "vouchers_duplicate": vouchers_duplicate,
        "vouchers_failed": vouchers_failed,
        "lines_processed": lines_processed,
        "accounts_balance_updated": accounts_updated,
        "total_rows": row_count,
        "errors": errors[:20],
        "error_count": len(errors)
    }


# ================== CATEGORIES IMPORT ==================

@router.post("/import/categories")
async def import_categories(
    file: UploadFile = File(...),
    organization_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Import product categories from Excel.
    Columns: CAT_ID, NAME
    """
    if current_user['role'] not in ['super_admin', 'admin']:
        raise HTTPException(status_code=403, detail="Only admins can import data")
    
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not available")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active
    
    created = 0
    updated = 0
    errors = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            cat_id = str(row[0]).strip() if row[0] else ''
            name = str(row[1]).strip() if row[1] else ''
            if not cat_id or not name:
                continue
            
            existing = await db.inventory_categories.find_one({
                'organization_id': organization_id,
                'cat_id': cat_id
            })
            
            if existing:
                await db.inventory_categories.update_one(
                    {'_id': existing['_id']},
                    {'$set': {'name': name}}
                )
                updated += 1
            else:
                await db.inventory_categories.insert_one({
                    'id': str(uuid.uuid4()),
                    'cat_id': cat_id,
                    'name': name,
                    'organization_id': organization_id,
                    'created_at': datetime.now(timezone.utc).isoformat()
                })
                created += 1
        except Exception as e:
            errors.append(str(e))
    
    wb.close()
    
    return {
        "message": "Categories imported successfully",
        "created": created,
        "updated": updated,
        "errors": errors[:20],
        "error_count": len(errors)
    }


# ================== REGIONS IMPORT ==================

@router.post("/import/regions")
async def import_regions(
    file: UploadFile = File(...),
    organization_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    """
    Import regions from Excel.
    Columns: REG_ID, NAME
    """
    if current_user['role'] not in ['super_admin', 'admin']:
        raise HTTPException(status_code=403, detail="Only admins can import data")
    
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not available")
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active
    
    created = 0
    updated = 0
    errors = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            reg_id = str(row[0]).strip() if row[0] else ''
            name = str(row[1]).strip() if row[1] else ''
            if not reg_id or not name:
                continue
            
            existing = await db.regions.find_one({
                'organization_id': organization_id,
                'reg_id': reg_id
            })
            
            if existing:
                await db.regions.update_one(
                    {'_id': existing['_id']},
                    {'$set': {'name': name}}
                )
                updated += 1
            else:
                await db.regions.insert_one({
                    'id': str(uuid.uuid4()),
                    'reg_id': reg_id,
                    'name': name,
                    'organization_id': organization_id,
                    'created_at': datetime.now(timezone.utc).isoformat()
                })
                created += 1
        except Exception as e:
            errors.append(str(e))
    
    wb.close()
    
    return {
        "message": "Regions imported successfully",
        "created": created,
        "updated": updated,
        "errors": errors[:20],
        "error_count": len(errors)
    }


# ================== INVENTORY ITEMS IMPORT ==================

@router.post("/import/inventory")
async def import_inventory(
    file: UploadFile = File(...),
    organization_id: str = Form(...),
    field_mapping: str = Form(None),  # JSON: {"item_code": 0, "name": 2, "category_id": 3, "supplier_id": 4, "package": 5, "pack_desc": 6, "price": 7, "cost": 8}
    current_user: dict = Depends(get_current_user)
):
    """Import inventory items with optional field mapping."""
    if current_user['role'] not in ['super_admin', 'admin']:
        raise HTTPException(status_code=403, detail="Only admins can import data")
    
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not available")
    
    # Parse field mapping
    mapping = None
    if field_mapping:
        try:
            mapping = json.loads(field_mapping)
        except json.JSONDecodeError:
            pass
    
    def get_col(row, field_name, default_idx):
        idx = mapping.get(field_name, default_idx) if mapping else default_idx
        if idx is None or idx < 0 or idx == '':
            return None
        idx = int(idx)
        return row[idx] if len(row) > idx and row[idx] is not None else None
    
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True)
    ws = wb.active
    
    # Pre-load categories for lookup
    categories = {}
    async for cat in db.inventory_categories.find({'organization_id': organization_id}, {'_id': 0}):
        categories[cat.get('cat_id', '')] = cat
    
    created = 0
    updated = 0
    errors = []
    items_batch = []
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            ic = get_col(row, 'item_code', 0)
            item_code = str(ic).strip() if ic else ''
            if not item_code:
                continue
            
            # Column mapping from ITEMS file using get_col
            desc_v = get_col(row, 'description', 1)
            description = str(desc_v).strip() if desc_v else ''
            name_v = get_col(row, 'name', 2)
            name_ar = str(name_v).strip() if name_v else ''
            cat_v = get_col(row, 'category_id', 3)
            cat_id = str(cat_v).strip() if cat_v else ''
            sup_v = get_col(row, 'supplier_id', 4)
            sup_id = str(sup_v).strip() if sup_v else ''
            pak_v = get_col(row, 'package', 5)
            pak = int(pak_v) if pak_v else 0
            pack_v = get_col(row, 'pack_desc', 6)
            pack_desc = str(pack_v).strip() if pack_v else ''
            price_v = get_col(row, 'price', 7)
            price = float(price_v or 0)
            cost_v = get_col(row, 'cost', 8)
            cost = float(cost_v or 0)
            
            # Get category name
            cat_name = categories.get(cat_id, {}).get('name', '')
            
            item_name = name_ar or description or f'Item {item_code}'
            
            item_doc = {
                'id': str(uuid.uuid4()),
                'item_code': item_code,
                'name': item_name,
                'name_ar': name_ar,
                'description': description,
                'category_id': cat_id,
                'category': cat_name,
                'category_name': cat_name,
                'supplier_id': sup_id,
                'supplier_name': '',
                'package': pak,
                'pack_description': pack_desc,
                'price': price,
                'cost': cost,
                'currency': 'USD',
                'unit': 'piece',
                'min_qty': 0,
                'on_hand_qty': 0,
                'is_active': True,
                'is_taxable': True,
                'organization_id': organization_id,
                'created_at': datetime.now(timezone.utc).isoformat()
            }
            
            items_batch.append(item_doc)
            
        except Exception as e:
            errors.append(f"Item {row[0] if row else '?'}: {str(e)}")
            if len(errors) > 50:
                break
    
    wb.close()
    
    # Check for existing items and insert/update
    if items_batch:
        existing_codes = set()
        existing = await db.inventory_items.find(
            {'organization_id': organization_id, 'item_code': {'$in': [i['item_code'] for i in items_batch]}},
            {'item_code': 1, '_id': 0}
        ).to_list(None)
        existing_codes = {e['item_code'] for e in existing}
        
        new_items = []
        for item in items_batch:
            if item['item_code'] in existing_codes:
                await db.inventory_items.update_one(
                    {'item_code': item['item_code'], 'organization_id': organization_id},
                    {'$set': {
                        'name': item['name'],
                        'name_ar': item['name_ar'],
                        'category_id': item['category_id'],
                        'category': item['category'],
                        'category_name': item['category_name'],
                        'supplier_id': item['supplier_id'],
                        'price': item['price'],
                        'cost': item['cost'],
                        'package': item['package'],
                        'pack_description': item['pack_description']
                    }}
                )
                updated += 1
            else:
                new_items.append(item)
                created += 1
        
        if new_items:
            for i in range(0, len(new_items), 500):
                batch = new_items[i:i+500]
                await db.inventory_items.insert_many(batch)
    
    return {
        "message": "Inventory imported successfully",
        "items_created": created,
        "items_updated": updated,
        "total_processed": len(items_batch),
        "errors": errors[:20],
        "error_count": len(errors)
    }


# ================== GET REGIONS & CATEGORIES ==================

@router.get("/regions")
async def get_regions(organization_id: str, current_user: dict = Depends(get_current_user)):
    """Get all regions for an organization"""
    regions = await db.regions.find(
        {'organization_id': organization_id}, {'_id': 0}
    ).sort('reg_id', 1).to_list(100)
    return regions

@router.get("/categories")
async def get_categories(organization_id: str, current_user: dict = Depends(get_current_user)):
    """Get all categories for an organization"""
    categories = await db.inventory_categories.find(
        {'organization_id': organization_id}, {'_id': 0}
    ).sort('cat_id', 1).to_list(100)
    return categories

